import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd


def process_meesho_gst_net_summary(
    sales_filepath, returns_filepath, output_filepath
):
  print("Processing Meesho TCS Sales & Return files...")

  # 1. Read Sales File
  df_sales = (
      pd.read_excel(sales_filepath)
      if sales_filepath.endswith((".xlsx", ".xls"))
      else pd.read_csv(sales_filepath)
  )

  # 2. Read Returns File
  df_returns = None
  if returns_filepath:
    df_returns = (
        pd.read_excel(returns_filepath)
        if returns_filepath.endswith((".xlsx", ".xls"))
        else pd.read_csv(returns_filepath)
    )

  # Column mappings explicitly specified:
  st_col = "end_customer_state_new"
  qty_col = "quantity"
  tax_col = "total_taxable_sale_value"

  # Fallback column detection in case of minor header variations
  if st_col not in df_sales.columns:
    st_col = [c for c in df_sales.columns if "state" in str(c).lower()][0]
  if qty_col not in df_sales.columns:
    qty_col = [
        c
        for c in df_sales.columns
        if "quantity" in str(c).lower() or "qty" in str(c).lower()
    ][0]
  if tax_col not in df_sales.columns:
    tax_col = [c for c in df_sales.columns if "taxable" in str(c).lower()][0]

  state_data = {}

  # Clean & Aggregate Sales Data
  df_sales[qty_col] = pd.to_numeric(
      df_sales[qty_col], errors="coerce"
  ).fillna(0)
  df_sales[tax_col] = pd.to_numeric(
      df_sales[tax_col], errors="coerce"
  ).fillna(0)
  df_sales[st_col] = df_sales[st_col].astype(str).str.strip().str.upper()

  for _, row in df_sales.iterrows():
    st_name = row[st_col]
    if not st_name or st_name in ["NAN", "NONE", ""]:
      continue
    if st_name not in state_data:
      state_data[st_name] = {
          "sales_pcs": 0,
          "sales_tax": 0.0,
          "ret_pcs": 0,
          "ret_tax": 0.0,
      }
    state_data[st_name]["sales_pcs"] += row[qty_col]
    state_data[st_name]["sales_tax"] += row[tax_col]

  # Clean & Aggregate Returns Data
  if df_returns is not None:
    r_st_col = (
        st_col
        if st_col in df_returns.columns
        else [c for c in df_returns.columns if "state" in str(c).lower()][0]
    )
    r_qty_col = (
        qty_col
        if qty_col in df_returns.columns
        else [
            c
            for c in df_returns.columns
            if "quantity" in str(c).lower() or "qty" in str(c).lower()
        ][0]
    )
    r_tax_col = (
        tax_col
        if tax_col in df_returns.columns
        else [c for c in df_returns.columns if "taxable" in str(c).lower()][0]
    )

    df_returns[r_qty_col] = pd.to_numeric(
        df_returns[r_qty_col], errors="coerce"
    ).fillna(0)
    df_returns[r_tax_col] = pd.to_numeric(
        df_returns[r_tax_col], errors="coerce"
    ).fillna(0)
    df_returns[r_st_col] = (
        df_returns[r_st_col].astype(str).str.strip().str.upper()
    )

    for _, row in df_returns.iterrows():
      st_name = row[r_st_col]
      if not st_name or st_name in ["NAN", "NONE", ""]:
        continue
      if st_name not in state_data:
        state_data[st_name] = {
            "sales_pcs": 0,
            "sales_tax": 0.0,
            "ret_pcs": 0,
            "ret_tax": 0.0,
        }
      state_data[st_name]["ret_pcs"] += row[r_qty_col]
      state_data[st_name]["ret_tax"] += row[r_tax_col]

  # Calculate Breakdown & Net Summary Rows
  summary_list = []
  for st_name in sorted(state_data.keys()):
    s_pcs = state_data[st_name]["sales_pcs"]
    s_tax = state_data[st_name]["sales_tax"]
    r_pcs = state_data[st_name]["ret_pcs"]
    r_tax = state_data[st_name]["ret_tax"]

    net_pcs = int(round(s_pcs - r_pcs))
    net_taxable = round(s_tax - r_tax, 2)
    avg_rate = round(net_taxable / net_pcs, 2) if net_pcs > 0 else 0.0

    summary_list.append({
        "STATE": st_name,
        "GROSS SALES TAXABLE": round(s_tax, 2),
        "RETURN TAXABLE": round(r_tax, 2),
        "NET PCS": net_pcs,
        "AVG RATE": avg_rate,
        "NET TAXABLE VALUE": net_taxable,
    })

  # Generate Formatted Excel Workbook
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "GSTR1 Net Taxable Summary"
  ws.views.sheetView[0].showGridLines = True

  fill_header = PatternFill(
      start_color="1F4E78", end_color="1F4E78", fill_type="solid"
  )
  fill_zebra = PatternFill(
      start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"
  )
  fill_net_col = PatternFill(
      start_color="E6F4EA", end_color="E6F4EA", fill_type="solid"
  )  # Soft green highlight
  fill_total = PatternFill(
      start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
  )

  font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
  font_data = Font(name="Segoe UI", size=10)
  font_total = Font(name="Segoe UI", size=10, bold=True)

  border_side = Side(style="thin", color="E0E0E0")
  border_data = Border(
      left=border_side, right=border_side, top=border_side, bottom=border_side
  )
  border_total = Border(
      top=Side(style="thin", color="000000"),
      bottom=Side(style="double", color="000000"),
  )

  indian_fmt = (
      "[>=10000000]₹##\\,##\\,##\\,##0.00;[>=100000]₹##\\,##\\,##0.00;₹##,##0.00"
  )

  headers = [
      "STATE",
      "GROSS SALES TAXABLE",
      "RETURN TAXABLE",
      "NET PCS",
      "AVG RATE",
      "TOTAL NET TAXABLE VALUE",
  ]
  for c_idx, h_text in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c_idx, value=h_text)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center")
  ws.row_dimensions[1].height = 26

  for idx, r in enumerate(summary_list, start=2):
    is_even = idx % 2 == 0

    ws.cell(row=idx, column=1, value=r["STATE"]).alignment = Alignment(
        horizontal="left", vertical="center"
    )

    c_gs = ws.cell(row=idx, column=2, value=r["GROSS SALES TAXABLE"])
    c_gs.number_format = indian_fmt
    c_gs.alignment = Alignment(horizontal="right", vertical="center")

    c_rt = ws.cell(row=idx, column=3, value=r["RETURN TAXABLE"])
    c_rt.number_format = indian_fmt
    c_rt.alignment = Alignment(horizontal="right", vertical="center")

    c_pcs = ws.cell(row=idx, column=4, value=r["NET PCS"])
    c_pcs.number_format = "#,##0"
    c_pcs.alignment = Alignment(horizontal="right", vertical="center")

    c_rate = ws.cell(row=idx, column=5, value=r["AVG RATE"])
    c_rate.number_format = indian_fmt
    c_rate.alignment = Alignment(horizontal="right", vertical="center")

    c_net = ws.cell(row=idx, column=6, value=r["NET TAXABLE VALUE"])
    c_net.number_format = indian_fmt
    c_net.alignment = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[idx].height = 20
    for col_c in range(1, 7):
      cell_obj = ws.cell(row=idx, column=col_c)
      cell_obj.font = font_data
      cell_obj.border = border_data
      if col_c == 6:
        cell_obj.fill = fill_net_col
      elif is_even:
        cell_obj.fill = fill_zebra

  # Total Summary Row
  tot_row = len(summary_list) + 2
  ws.row_dimensions[tot_row].height = 24
  ws.cell(row=tot_row, column=1, value="Total").alignment = Alignment(
      horizontal="left", vertical="center"
  )

  ws.cell(row=tot_row, column=2, value=f"=SUM(B2:B{tot_row-1})").number_format = (
      indian_fmt
  )
  ws.cell(row=tot_row, column=2).alignment = Alignment(
      horizontal="right", vertical="center"
  )

  ws.cell(row=tot_row, column=3, value=f"=SUM(C2:C{tot_row-1})").number_format = (
      indian_fmt
  )
  ws.cell(row=tot_row, column=3).alignment = Alignment(
      horizontal="right", vertical="center"
  )

  ws.cell(row=tot_row, column=4, value=f"=SUM(D2:D{tot_row-1})").number_format = (
      "#,##0"
  )
  ws.cell(row=tot_row, column=4).alignment = Alignment(
      horizontal="right", vertical="center"
  )

  ws.cell(row=tot_row, column=5, value="-").alignment = Alignment(
      horizontal="center", vertical="center"
  )

  ws.cell(row=tot_row, column=6, value=f"=SUM(F2:F{tot_row-1})").number_format = (
      indian_fmt
  )
  ws.cell(row=tot_row, column=6).alignment = Alignment(
      horizontal="right", vertical="center"
  )

  for col_c in range(1, 7):
    cell_obj = ws.cell(row=tot_row, column=col_c)
    cell_obj.font = font_total
    cell_obj.border = border_total
    if col_c != 6:
      cell_obj.fill = fill_total

  ws.column_dimensions["A"].width = 36
  ws.column_dimensions["B"].width = 24
  ws.column_dimensions["C"].width = 20
  ws.column_dimensions["D"].width = 16
  ws.column_dimensions["E"].width = 18
  ws.column_dimensions["F"].width = 28

  wb.save(output_filepath)
  print(f"File created successfully: {output_filepath}")


# Execution
process_meesho_gst_net_summary(
    "tcs_sales.xlsx", "tcs_sales_return.xlsx", "Meesho_Final_Net_GSTR1.xlsx"
)
 
